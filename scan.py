import argparse
import base64
import json
import os
import re
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from dotenv import load_dotenv
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

EXCEL_COLUMNS = ["Date", "Vendor", "Amount", "CC #", "COA", "Location"]
DEFAULT_OUTPUT = Path("receipt_log.xlsx")
CAPTURE_DIR = Path("captures")
AI_LOG_DIR = Path("ai_logs")
CAMERA_BACKENDS = {
    "any": 0,
    "dshow": cv2.CAP_DSHOW,
    "msmf": cv2.CAP_MSMF,
}


@dataclass
class Detection:
    corners: np.ndarray
    area_ratio: float


def order_points(points: np.ndarray) -> np.ndarray:
    rect = np.zeros((4, 2), dtype="float32")
    pts = points.reshape(4, 2).astype("float32")

    sums = pts.sum(axis=1)
    diffs = np.diff(pts, axis=1)

    rect[0] = pts[np.argmin(sums)]
    rect[2] = pts[np.argmax(sums)]
    rect[1] = pts[np.argmin(diffs)]
    rect[3] = pts[np.argmax(diffs)]
    return rect


def find_receipt(frame: np.ndarray, min_area_ratio: float = 0.12) -> Detection | None:
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(blurred, 60, 180)
    edges = cv2.dilate(edges, np.ones((3, 3), np.uint8), iterations=1)

    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    frame_area = frame.shape[0] * frame.shape[1]

    candidates: list[Detection] = []
    for contour in contours:
        perimeter = cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, 0.02 * perimeter, True)

        if len(approx) != 4 or not cv2.isContourConvex(approx):
            continue

        area = cv2.contourArea(approx)
        area_ratio = area / frame_area
        if area_ratio >= min_area_ratio:
            candidates.append(Detection(order_points(approx), area_ratio))

    if not candidates:
        return None

    return max(candidates, key=lambda detection: detection.area_ratio)


def corners_are_stable(current: np.ndarray, previous: np.ndarray | None, tolerance: float) -> bool:
    if previous is None:
        return False
    return float(np.mean(np.linalg.norm(current - previous, axis=1))) <= tolerance


def focus_score(image: np.ndarray) -> float:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    return float(cv2.Laplacian(gray, cv2.CV_64F).var())


def crop_to_bounds(frame: np.ndarray, corners: np.ndarray, padding: int = 8) -> np.ndarray:
    height, width = frame.shape[:2]
    points = corners.reshape(4, 2)
    x_min = max(0, int(np.min(points[:, 0])) - padding)
    y_min = max(0, int(np.min(points[:, 1])) - padding)
    x_max = min(width, int(np.max(points[:, 0])) + padding)
    y_max = min(height, int(np.max(points[:, 1])) + padding)
    return frame[y_min:y_max, x_min:x_max]


def warp_receipt(frame: np.ndarray, corners: np.ndarray) -> np.ndarray:
    top_left, top_right, bottom_right, bottom_left = corners

    width_a = np.linalg.norm(bottom_right - bottom_left)
    width_b = np.linalg.norm(top_right - top_left)
    max_width = int(max(width_a, width_b))

    height_a = np.linalg.norm(top_right - bottom_right)
    height_b = np.linalg.norm(top_left - bottom_left)
    max_height = int(max(height_a, height_b))

    destination = np.array(
        [[0, 0], [max_width - 1, 0], [max_width - 1, max_height - 1], [0, max_height - 1]],
        dtype="float32",
    )
    matrix = cv2.getPerspectiveTransform(corners, destination)
    warped = cv2.warpPerspective(frame, matrix, (max_width, max_height))
    return enhance_for_ai(warped)


def enhance_for_ai(image: np.ndarray) -> np.ndarray:
    height, width = image.shape[:2]
    scale = max(1.0, 1100 / max(width, height))
    if scale > 1:
        image = cv2.resize(image, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)

    lab = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
    lightness, a_channel, b_channel = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    lightness = clahe.apply(lightness)
    enhanced = cv2.merge((lightness, a_channel, b_channel))
    return cv2.cvtColor(enhanced, cv2.COLOR_LAB2BGR)


def make_debug_ocr_image(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, h=10)
    return cv2.adaptiveThreshold(
        denoised,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31,
        8,
    )


def save_capture(image: np.ndarray) -> Path:
    CAPTURE_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = CAPTURE_DIR / f"receipt_{timestamp}.png"
    cv2.imwrite(str(path), image)
    return path


def save_debug_ocr_image(image: np.ndarray, capture_path: Path) -> Path:
    debug_path = capture_path.with_name(f"{capture_path.stem}_ocr_debug.png")
    cv2.imwrite(str(debug_path), make_debug_ocr_image(image))
    return debug_path


def encode_image(path: Path) -> str:
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def extract_json(text: str) -> dict[str, Any]:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, flags=re.DOTALL)
        if not match:
            raise
        return json.loads(match.group(0))


def analyze_receipt(image_path: Path, model: str) -> dict[str, Any]:
    try:
        from openai import OpenAI
    except ImportError as exc:
        raise RuntimeError("Install dependencies first: pip install -r requirements.txt") from exc

    client = OpenAI()
    image_b64 = encode_image(image_path)
    prompt = (
        "Read the receipt image carefully and extract one expense row. "
        "Return JSON only with these keys: date, vendor, amount, cc_number, coa, location, confidence, notes. "
        "Rules: date must be YYYY-MM-DD if visible, otherwise null. "
        "vendor must be the merchant name exactly as printed, correcting obvious OCR spacing but not inventing. "
        "amount must be the final total charged, not subtotal, tax, item price, rewards, balance, or cash tendered. "
        "cc_number must be the last 4 card digits, cash, or null. "
        "coa must be one short accounting category; use auto for vehicle parts, fuel, tires, repair, or car wash. "
        "location must be the city/store location only if printed on the receipt, otherwise null. "
        "confidence must be high, medium, or low. "
        "notes should briefly mention any fields that were hard to read."
    )

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": prompt},
                    {
                        "type": "input_image",
                        "image_url": f"data:image/png;base64,{image_b64}",
                    },
                ],
            }
        ],
    )
    raw_text = response.output_text
    save_ai_log(image_path, raw_text)
    return normalize_receipt_data(extract_json(raw_text))


def save_ai_log(image_path: Path, raw_text: str) -> Path:
    AI_LOG_DIR.mkdir(exist_ok=True)
    path = AI_LOG_DIR / f"{image_path.stem}.json"
    path.write_text(raw_text, encoding="utf-8")
    return path


def normalize_receipt_data(data: dict[str, Any]) -> dict[str, Any]:
    amount = data.get("amount")
    if isinstance(amount, str):
        amount = amount.replace("$", "").replace(",", "").strip()
    try:
        amount = float(amount) if amount not in (None, "") else None
    except (TypeError, ValueError):
        amount = None

    date_value = data.get("date")
    if isinstance(date_value, str):
        date_value = parse_date(date_value.strip())

    cc_number = data.get("cc_number")
    if cc_number is not None:
        cc_number = str(cc_number).strip()

    normalized = {
        "Date": date_value or None,
        "Vendor": data.get("vendor") or None,
        "Amount": amount,
        "CC #": cc_number or None,
        "COA": data.get("coa") or "uncategorized",
        "Location": data.get("location") or None,
    }
    normalized["_confidence"] = data.get("confidence") or None
    normalized["_notes"] = data.get("notes") or None
    return normalized


def parse_date(value: str) -> date | str | None:
    if not value:
        return None

    for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue

    return value


def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active
        existing_headers = [sheet.cell(row=1, column=i + 1).value for i in range(len(EXCEL_COLUMNS))]
        if existing_headers != EXCEL_COLUMNS:
            raise ValueError(f"{path} exists but does not use the expected headers: {EXCEL_COLUMNS}")
        return workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Receipts"
    sheet.append(EXCEL_COLUMNS)
    style_sheet(sheet)
    return workbook


def style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column_index, header in enumerate(EXCEL_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
        sheet.column_dimensions[get_column_letter(column_index)].width = [12, 24, 12, 10, 14, 18][column_index - 1]

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}1"


def append_to_workbook(path: Path, receipt: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = ensure_workbook(path)
    sheet = workbook.active
    sheet.append([receipt.get(column) for column in EXCEL_COLUMNS])

    row = sheet.max_row
    sheet.cell(row=row, column=1).number_format = "m/d/yyyy"
    sheet.cell(row=row, column=3).number_format = '"$"#,##0.00'
    sheet.cell(row=row, column=4).number_format = "@"
    style_sheet(sheet)
    workbook.save(path)


def display_receipt(receipt: dict[str, Any]) -> None:
    print("\nExtracted receipt data:")
    for column in EXCEL_COLUMNS:
        print(f"  {column}: {receipt.get(column)}")
    if receipt.get("_confidence"):
        print(f"  Confidence: {receipt.get('_confidence')}")
    if receipt.get("_notes"):
        print(f"  Notes: {receipt.get('_notes')}")


def review_receipt(receipt: dict[str, Any]) -> dict[str, Any] | None:
    display_receipt(receipt)
    print("\nPress Enter to accept, type s to skip, or paste corrected JSON for these keys:")
    print('{"date":"2025-01-02","vendor":"Costco","amount":65.15,"cc_number":"5121","coa":"auto","location":"Fullerton"}')

    while True:
        user_input = input("> ").strip()

        if not user_input:
            return receipt
        if user_input.lower() in {"s", "skip"}:
            return None

        try:
            corrected = normalize_receipt_data(extract_json(user_input))
        except json.JSONDecodeError as exc:
            print(f"That correction was not valid JSON: {exc.msg} at character {exc.pos}.")
            print('Use double quotes around text, like {"vendor":"Costco","coa":"auto"}, or type s to skip.')
            continue

        display_receipt(corrected)
        print("Press Enter to save this corrected row, type e to edit again, or type s to skip.")
        confirmation = input("> ").strip().lower()
        if confirmation in {"s", "skip"}:
            return None
        if confirmation in {"e", "edit"}:
            print("Paste corrected JSON again:")
            continue
        return corrected


def draw_overlay(
    frame: np.ndarray,
    detection: Detection | None,
    stable_count: int,
    required_count: int,
    sharpness: float | None,
    min_sharpness: float,
) -> None:
    if detection is not None:
        cv2.drawContours(frame, [detection.corners.astype(int)], -1, (0, 255, 0), 3)

    receipt_status = "Receipt: found" if detection is not None else "Receipt: not found"
    status = f"{receipt_status} | stability: {stable_count}/{required_count}"
    focus = "Focus: no receipt"
    focus_color = (255, 255, 255)
    if sharpness is not None:
        focus_ready = sharpness >= min_sharpness
        focus = f"Focus: {sharpness:.0f}/{min_sharpness:.0f} {'OK' if focus_ready else 'WAIT'}"
        focus_color = (20, 220, 20) if focus_ready else (0, 200, 255)

    controls = "q quit | c force capture"
    cv2.putText(frame, status, (18, 32), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (20, 220, 20), 2)
    cv2.putText(frame, focus, (18, 64), cv2.FONT_HERSHEY_SIMPLEX, 0.7, focus_color, 2)
    cv2.putText(frame, controls, (18, 96), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)


def process_capture(
    frame: np.ndarray,
    corners: np.ndarray,
    output: Path,
    model: str,
    dry_run: bool,
    review: bool,
    save_debug: bool,
) -> None:
    receipt_image = warp_receipt(frame, corners)
    image_path = save_capture(receipt_image)
    print(f"Captured receipt image: {image_path}")
    if save_debug:
        debug_path = save_debug_ocr_image(receipt_image, image_path)
        print(f"Saved OCR debug image: {debug_path}")

    if dry_run:
        print("Dry run enabled, skipping AI extraction and Excel append.")
        return

    receipt = analyze_receipt(image_path, model)
    if review:
        receipt = review_receipt(receipt)
        if receipt is None:
            print("Skipped Excel append for this capture.")
            return

    append_to_workbook(output, receipt)
    print(f"Added row to {output}: {receipt}")


def open_camera(camera_index: int, backend: str) -> Any:
    backend_value = CAMERA_BACKENDS[backend]
    if backend == "any":
        return cv2.VideoCapture(camera_index)
    return cv2.VideoCapture(camera_index, backend_value)


def run_scanner(args: argparse.Namespace) -> None:
    if not args.dry_run and not os.getenv("OPENAI_API_KEY"):
        raise RuntimeError("Set OPENAI_API_KEY before running AI extraction, or use --dry-run.")

    capture = open_camera(args.camera, args.backend)
    if not capture.isOpened():
        raise RuntimeError(f"Could not open camera index {args.camera}")
    if args.width > 0:
        capture.set(cv2.CAP_PROP_FRAME_WIDTH, args.width)
    if args.height > 0:
        capture.set(cv2.CAP_PROP_FRAME_HEIGHT, args.height)
    if not args.no_autofocus:
        capture.set(cv2.CAP_PROP_AUTOFOCUS, 1)

    previous_corners: np.ndarray | None = None
    stable_count = 0
    sharp_count = 0
    failed_reads = 0
    last_capture_time = 0.0

    try:
        while True:
            ok, frame = capture.read()
            if not ok:
                failed_reads += 1
                print(f"Camera frame could not be read; retrying ({failed_reads}/{args.max_read_failures}).")
                if failed_reads >= args.max_read_failures:
                    raise RuntimeError(
                        "Camera stopped returning frames. Try: python scan.py --backend dshow --width 0 --height 0"
                    )
                time.sleep(0.1)
                continue
            failed_reads = 0

            detection = find_receipt(frame, args.min_area_ratio)
            if detection and corners_are_stable(detection.corners, previous_corners, args.stability_tolerance):
                stable_count += 1
            else:
                stable_count = 1 if detection else 0

            sharpness = None
            if detection:
                receipt_region = crop_to_bounds(frame, detection.corners)
                sharpness = focus_score(receipt_region)
                if sharpness >= args.min_sharpness:
                    sharp_count += 1
                else:
                    sharp_count = 0
            else:
                sharp_count = 0

            previous_corners = detection.corners if detection else None
            draw_overlay(frame, detection, stable_count, args.stable_frames, sharpness, args.min_sharpness)
            cv2.imshow("Receipt Scanner", frame)

            key = cv2.waitKey(1) & 0xFF
            ready_to_capture = (
                detection is not None
                and stable_count >= args.stable_frames
                and sharp_count >= args.sharp_frames
                and time.time() - last_capture_time >= args.cooldown_seconds
            )

            if key == ord("q"):
                break
            if key == ord("c") and detection is not None:
                ready_to_capture = True

            if ready_to_capture and detection is not None:
                process_capture(
                    frame,
                    detection.corners,
                    args.output,
                    args.model,
                    args.dry_run,
                    not args.no_review,
                    args.save_debug,
                )
                last_capture_time = time.time()
                stable_count = 0
                sharp_count = 0
    finally:
        capture.release()
        cv2.destroyAllWindows()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Detect receipts from a live camera and log AI-extracted data to Excel.")
    parser.add_argument("--camera", type=int, default=0, help="OpenCV camera index.")
    parser.add_argument("--backend", choices=CAMERA_BACKENDS.keys(), default="dshow", help="OpenCV camera backend.")
    parser.add_argument("--width", type=int, default=1280, help="Requested camera capture width. Use 0 for camera default.")
    parser.add_argument("--height", type=int, default=720, help="Requested camera capture height. Use 0 for camera default.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Excel workbook path.")
    parser.add_argument("--model", default="gpt-4.1-mini", help="OpenAI vision-capable model.")
    parser.add_argument("--dry-run", action="store_true", help="Capture receipt images without calling AI or writing Excel rows.")
    parser.add_argument("--no-review", action="store_true", help="Append AI extraction to Excel without asking for confirmation.")
    parser.add_argument("--save-debug", action="store_true", help="Also save a black-and-white OCR debug image beside each capture.")
    parser.add_argument("--no-autofocus", action="store_true", help="Do not ask the camera to enable autofocus.")
    parser.add_argument("--stable-frames", type=int, default=12, help="Frames the receipt must remain stable before capture.")
    parser.add_argument("--sharp-frames", type=int, default=6, help="Focused frames required before automatic capture.")
    parser.add_argument("--min-sharpness", type=float, default=120.0, help="Minimum Laplacian focus score before automatic capture.")
    parser.add_argument("--stability-tolerance", type=float, default=12.0, help="Average corner movement allowed between frames.")
    parser.add_argument("--cooldown-seconds", type=float, default=4.0, help="Minimum seconds between automatic captures.")
    parser.add_argument("--min-area-ratio", type=float, default=0.05, help="Minimum receipt size as a fraction of frame area.")
    parser.add_argument("--max-read-failures", type=int, default=20, help="Failed camera reads before exiting.")
    return parser.parse_args()


if __name__ == "__main__":
    run_scanner(parse_args())
